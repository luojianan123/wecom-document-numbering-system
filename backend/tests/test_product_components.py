from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from .conftest import login


def test_component_project_full_numbering_flow(client: TestClient) -> None:
    csrf = login(client, "user", "component-user")
    missing = client.get("/api/component-codes/projects/2468")
    assert missing.status_code == 404

    created = client.post(
        "/api/component-codes/projects",
        json={
            "project_code": "2468",
            "machine_name": "星海处理机",
            "is_prototype": False,
        },
        headers={"X-CSRF-Token": csrf},
    )
    assert created.status_code == 200, created.text
    project = created.json()
    machine = project["nodes"][0]
    assert machine["code"] == "GH2468-XHCLJ-00-G-1.00"

    component = client.post(
        f"/api/component-codes/projects/{project['id']}/nodes",
        json={
            "parent_id": machine["id"],
            "kind": "component",
            "name": "主控部组件",
            "is_prototype": True,
        },
        headers={"X-CSRF-Token": csrf},
    )
    assert component.status_code == 200, component.text
    component_node = component.json()
    assert component_node["code"] == "GH2468-XHCLJ-01-00-Z-2.00"

    structure = client.post(
        f"/api/component-codes/projects/{project['id']}/nodes",
        json={
            "parent_id": component_node["id"],
            "kind": "structure",
            "name": "机箱结构",
            "is_prototype": False,
        },
        headers={"X-CSRF-Token": csrf},
    ).json()
    hardware = client.post(
        f"/api/component-codes/projects/{project['id']}/nodes",
        json={
            "parent_id": component_node["id"],
            "kind": "hardware",
            "name": "主控板",
            "is_prototype": False,
        },
        headers={"X-CSRF-Token": csrf},
    ).json()
    assert structure["code"] == "GH2468-XHCLJ-01-01-G-1.00"
    assert hardware["code"] == "GH2468-XHCLJ-01-10-G-1.00"

    part = client.post(
        f"/api/component-codes/projects/{project['id']}/nodes",
        json={
            "parent_id": hardware["id"],
            "kind": "part",
            "name": "处理器",
            "is_prototype": False,
        },
        headers={"X-CSRF-Token": csrf},
    )
    assert part.status_code == 200, part.text
    assert part.json()["code"] == "GH2468-XHCLJ-01-10-00-G-1.00"

    software = client.post(
        f"/api/component-codes/projects/{project['id']}/nodes",
        json={
            "parent_id": component_node["id"],
            "kind": "software",
            "name": "控制软件",
            "is_prototype": False,
        },
        headers={"X-CSRF-Token": csrf},
    )
    assert software.status_code == 400
    assert "规则尚未配置" in software.json()["detail"]

    other = client.post(
        f"/api/component-codes/projects/{project['id']}/nodes",
        json={
            "parent_id": component_node["id"],
            "kind": "other",
            "name": "其他组成",
            "is_prototype": False,
        },
        headers={"X-CSRF-Token": csrf},
    )
    assert other.status_code == 200, other.text
    assert other.json()["code"] == "GH2468-XHCLJ-01-20-G-1.00"

    other_prototype = client.post(
        f"/api/component-codes/projects/{project['id']}/nodes",
        json={
            "parent_id": component_node["id"],
            "kind": "other",
            "name": "其他正样组成",
            "is_prototype": True,
        },
        headers={"X-CSRF-Token": csrf},
    )
    assert other_prototype.status_code == 200, other_prototype.text
    assert other_prototype.json()["code"] == "GH2468-XHCLJ-01-21-Z-2.00"


def test_structure_sequence_moves_to_thirty_after_ten(client: TestClient) -> None:
    csrf = login(client, "user", "component-structure-sequence")
    project = client.post(
        "/api/component-codes/projects",
        json={"project_code": "2469", "machine_name": "结构序号验证机"},
        headers={"X-CSRF-Token": csrf},
    ).json()
    machine = project["nodes"][0]
    component = client.post(
        f"/api/component-codes/projects/{project['id']}/nodes",
        json={
            "parent_id": machine["id"],
            "kind": "component",
            "name": "结构序号验证部组件",
        },
        headers={"X-CSRF-Token": csrf},
    ).json()

    structures = []
    for index in range(11):
        structure = client.post(
            f"/api/component-codes/projects/{project['id']}/nodes",
            json={
                "parent_id": component["id"],
                "kind": "structure",
                "name": f"结构{index + 1}",
            },
            headers={"X-CSRF-Token": csrf},
        )
        assert structure.status_code == 200, structure.text
        structures.append(structure.json())

    assert [node["sequence"] for node in structures] == [*range(1, 11), 30]
    assert structures[-1]["code"] == "GH2469-JGXHYZ-01-30-G-1.00"


def test_admin_lists_component_projects_with_creators_and_claims(
    client: TestClient,
) -> None:
    user_csrf = login(client, "user", "component-list-creator")
    project = client.post(
        "/api/component-codes/projects",
        json={"project_code": "7401", "machine_name": "列表验证机"},
        headers={"X-CSRF-Token": user_csrf},
    ).json()
    machine = project["nodes"][0]
    client.post(
        f"/api/component-codes/nodes/{machine['id']}/claim",
        headers={"X-CSRF-Token": user_csrf},
    )

    admin_csrf = login(client, "admin", "component-list-admin")
    listed = client.get("/api/component-codes/projects/admin/list")
    assert listed.status_code == 200, listed.text
    summary = next(item for item in listed.json() if item["project_code"] == "7401")
    assert summary["created_by_name"] == "普通用户"
    assert summary["machine_count"] == 1
    assert summary["node_count"] == 1
    assert summary["claim_count"] == 1

    detail = client.get("/api/component-codes/projects/7401")
    assert detail.status_code == 200, detail.text
    assert detail.json()["created_by_name"] == "普通用户"
    assert detail.json()["nodes"][0]["created_by_name"] == "普通用户"
    assert detail.json()["nodes"][0]["claims"][0]["claimant_name"] == "普通用户"
    assert admin_csrf


def test_component_edit_claim_export_and_recursive_delete(client: TestClient) -> None:
    csrf = login(client, "user", "component-editor")
    project = client.post(
        "/api/component-codes/projects",
        json={"project_code": "1357", "machine_name": "测试机", "is_prototype": True},
        headers={"X-CSRF-Token": csrf},
    ).json()
    machine = project["nodes"][0]
    component = client.post(
        f"/api/component-codes/projects/{project['id']}/nodes",
        json={
            "parent_id": machine["id"],
            "kind": "component",
            "name": "测试组件",
            "is_prototype": False,
        },
        headers={"X-CSRF-Token": csrf},
    ).json()

    edited = client.post(
        f"/api/component-codes/projects/{project['id']}/nodes/{component['id']}",
        json={"name": "修正组件", "code": "GH1357-CSJ-01-00-G-1.00"},
        headers={"X-CSRF-Token": csrf},
    )
    assert edited.status_code == 200
    assert edited.json()["name"] == "修正组件"

    claim = client.post(
        f"/api/component-codes/nodes/{component['id']}/claim",
        headers={"X-CSRF-Token": csrf},
    )
    assert claim.status_code == 200
    assert claim.json()["claimant_name"] == "普通用户"

    exported = client.get(f"/api/component-codes/projects/{project['id']}/export")
    assert exported.status_code == 200
    workbook = load_workbook(BytesIO(exported.content), read_only=True)
    rows = list(workbook.active.iter_rows(values_only=True))
    assert workbook.sheetnames == ["测试机"]
    assert rows[0][0] == "项目号：1357"
    assert rows[1][0] == "整机"
    assert rows[1][3] == "部组件"
    assert rows[2][:6] == ("名称", "编号", "阶段", "名称", "编号", "阶段")
    assert any(row[3] == "修正组件" for row in rows)
    workbook.close()

    deleted = client.post(
        f"/api/component-codes/projects/{project['id']}/bulk-delete",
        json={"node_ids": [machine["id"]]},
        headers={"X-CSRF-Token": csrf},
    )
    assert deleted.status_code == 204
    detail = client.get("/api/component-codes/projects/1357").json()
    assert detail["nodes"] == []


def test_component_export_uses_one_tree_sheet_per_machine(client: TestClient) -> None:
    csrf = login(client, "user", "component-multi-machine")
    project = client.post(
        "/api/component-codes/projects/tree/generate",
        json={
            "project_code": "3579",
            "nodes": [
                {"client_id": "m1", "kind": "machine", "name": "甲整机"},
                {
                    "client_id": "c1",
                    "parent_client_id": "m1",
                    "kind": "component",
                    "name": "甲组件",
                },
                {"client_id": "m2", "kind": "machine", "name": "乙整机"},
                {
                    "client_id": "c2",
                    "parent_client_id": "m2",
                    "kind": "component",
                    "name": "乙组件",
                },
            ],
        },
        headers={"X-CSRF-Token": csrf},
    ).json()

    exported = client.get(f"/api/component-codes/projects/{project['id']}/export")
    assert exported.status_code == 200
    workbook = load_workbook(BytesIO(exported.content), read_only=True)
    assert workbook.sheetnames == ["甲整机", "乙整机"]
    first_rows = list(workbook["甲整机"].iter_rows(values_only=True))
    second_rows = list(workbook["乙整机"].iter_rows(values_only=True))
    assert any(row[0] == "甲整机" and row[3] == "甲组件" for row in first_rows)
    assert any(row[0] == "乙整机" and row[3] == "乙组件" for row in second_rows)
    workbook.close()


def test_component_delete_renumbers_following_tree(client: TestClient) -> None:
    csrf = login(client, "user", "component-delete-renumber")
    project = client.post(
        "/api/component-codes/projects/tree/generate",
        json={
            "project_code": "4680",
            "nodes": [
                {"client_id": "machine", "kind": "machine", "name": "重排验证机"},
                {
                    "client_id": "component-1",
                    "parent_client_id": "machine",
                    "kind": "component",
                    "name": "待删除组件",
                },
                {
                    "client_id": "component-2",
                    "parent_client_id": "machine",
                    "kind": "component",
                    "name": "保留组件",
                },
                {
                    "client_id": "hardware",
                    "parent_client_id": "component-2",
                    "kind": "hardware",
                    "name": "保留硬件",
                },
                {
                    "client_id": "part",
                    "parent_client_id": "hardware",
                    "kind": "part",
                    "name": "保留零件",
                },
            ],
        },
        headers={"X-CSRF-Token": csrf},
    ).json()
    by_name = {node["name"]: node for node in project["nodes"]}
    assert by_name["保留组件"]["code"] == "GH4680-ZPYZJ-02-00-G-1.00"
    assert by_name["保留硬件"]["code"] == "GH4680-ZPYZJ-02-10-G-1.00"
    assert by_name["保留零件"]["code"] == "GH4680-ZPYZJ-02-10-00-G-1.00"

    deleted = client.post(
        f"/api/component-codes/projects/{project['id']}/bulk-delete",
        json={"node_ids": [by_name["待删除组件"]["id"]]},
        headers={"X-CSRF-Token": csrf},
    )
    assert deleted.status_code == 204, deleted.text
    before_renumber = client.get("/api/component-codes/projects/4680").json()
    stale = {node["name"]: node["code"] for node in before_renumber["nodes"]}
    assert stale["保留组件"] == "GH4680-ZPYZJ-02-00-G-1.00"

    generated = client.post(
        f"/api/component-codes/projects/{project['id']}/renumber",
        headers={"X-CSRF-Token": csrf},
    )
    assert generated.status_code == 200, generated.text
    refreshed = client.get("/api/component-codes/projects/4680").json()
    renumbered = {node["name"]: node["code"] for node in refreshed["nodes"]}
    assert "待删除组件" not in renumbered
    assert renumbered["保留组件"] == "GH4680-ZPYZJ-01-00-G-1.00"
    assert renumbered["保留硬件"] == "GH4680-ZPYZJ-01-10-G-1.00"
    assert renumbered["保留零件"] == "GH4680-ZPYZJ-01-10-00-G-1.00"


def test_component_delete_first_part_resets_remaining_part_to_zero(
    client: TestClient,
) -> None:
    csrf = login(client, "user", "component-delete-part")
    project = client.post(
        "/api/component-codes/projects/tree/generate",
        json={
            "project_code": "1284",
            "nodes": [
                {"client_id": "machine", "kind": "machine", "name": "控制机"},
                {
                    "client_id": "component",
                    "parent_client_id": "machine",
                    "kind": "component",
                    "name": "控制单元板PCBA",
                },
                {
                    "client_id": "hardware",
                    "parent_client_id": "component",
                    "kind": "hardware",
                    "name": "控制单元板",
                },
                {
                    "client_id": "part-1",
                    "parent_client_id": "hardware",
                    "kind": "part",
                    "name": "控制单元板PCB",
                },
                {
                    "client_id": "part-2",
                    "parent_client_id": "hardware",
                    "kind": "part",
                    "name": "控制单元板PCB设计",
                },
            ],
        },
        headers={"X-CSRF-Token": csrf},
    ).json()
    by_name = {node["name"]: node for node in project["nodes"]}
    assert by_name["控制单元板PCB设计"]["code"].endswith("-01-G-1.00")

    deleted = client.post(
        f"/api/component-codes/projects/{project['id']}/bulk-delete",
        json={"node_ids": [by_name["控制单元板PCB"]["id"]]},
        headers={"X-CSRF-Token": csrf},
    )
    assert deleted.status_code == 204, deleted.text
    before_renumber = client.get("/api/component-codes/projects/1284").json()
    stale_part = next(
        node
        for node in before_renumber["nodes"]
        if node["name"] == "控制单元板PCB设计"
    )
    assert stale_part["code"].endswith("-01-G-1.00")

    generated = client.post(
        f"/api/component-codes/projects/{project['id']}/renumber",
        headers={"X-CSRF-Token": csrf},
    )
    assert generated.status_code == 200, generated.text
    refreshed = client.get("/api/component-codes/projects/1284").json()
    remaining = next(
        node for node in refreshed["nodes"] if node["name"] == "控制单元板PCB设计"
    )
    assert remaining["code"].endswith("-00-G-1.00")
    assert remaining["sequence"] == 0


def test_component_manual_edit_validates_level_and_updates_metadata(
    client: TestClient,
) -> None:
    csrf = login(client, "user", "component-validator")
    project = client.post(
        "/api/component-codes/projects",
        json={"project_code": "8642", "machine_name": "验证机", "is_prototype": False},
        headers={"X-CSRF-Token": csrf},
    ).json()
    machine = project["nodes"][0]
    component = client.post(
        f"/api/component-codes/projects/{project['id']}/nodes",
        json={
            "parent_id": machine["id"],
            "kind": "component",
            "name": "验证组件",
            "is_prototype": False,
        },
        headers={"X-CSRF-Token": csrf},
    ).json()
    hardware = client.post(
        f"/api/component-codes/projects/{project['id']}/nodes",
        json={
            "parent_id": component["id"],
            "kind": "hardware",
            "name": "验证硬件",
            "is_prototype": False,
        },
        headers={"X-CSRF-Token": csrf},
    ).json()

    invalid = client.post(
        f"/api/component-codes/projects/{project['id']}/nodes/{hardware['id']}",
        json={"name": "验证硬件", "code": "GH8642-YZJ-01-01-G-1.00"},
        headers={"X-CSRF-Token": csrf},
    )
    assert invalid.status_code == 400
    assert "硬件序列号必须从10开始" in invalid.json()["detail"]

    wrong_parent = client.post(
        f"/api/component-codes/projects/{project['id']}/nodes/{hardware['id']}",
        json={"name": "验证硬件", "code": "GH8642-OTHER-01-12-Z-2.00"},
        headers={"X-CSRF-Token": csrf},
    )
    assert wrong_parent.status_code == 400
    assert "上级继承段" in wrong_parent.json()["detail"]

    edited = client.post(
        f"/api/component-codes/projects/{project['id']}/nodes/{hardware['id']}",
        json={"name": "验证硬件正样", "code": "gh8642-yzj-01-12-z-2.00"},
        headers={"X-CSRF-Token": csrf},
    )
    assert edited.status_code == 200, edited.text
    assert edited.json()["code"] == "GH8642-YZJ-01-12-Z-2.00"
    assert edited.json()["stage"] == "Z"
    assert edited.json()["sequence"] == 12

    wrong_version = client.post(
        f"/api/component-codes/projects/{project['id']}/nodes/{hardware['id']}",
        json={"name": "验证硬件正样", "code": "GH8642-YZJ-01-12-Z-1.00"},
        headers={"X-CSRF-Token": csrf},
    )
    assert wrong_version.status_code == 400
    assert "Z阶段的版本号必须为2.00" in wrong_version.json()["detail"]


def test_generate_complete_component_tree_in_one_batch(client: TestClient) -> None:
    csrf = login(client, "user", "component-tree-user")
    generated = client.post(
        "/api/component-codes/projects/tree/generate",
        json={
            "project_code": "9753",
            "nodes": [
                {
                    "client_id": "part-1",
                    "parent_client_id": "hardware-1",
                    "kind": "part",
                    "name": "处理器",
                    "is_prototype": False,
                },
                {
                    "client_id": "component-1",
                    "parent_client_id": "machine-1",
                    "kind": "component",
                    "name": "主控部组件",
                    "is_prototype": False,
                },
                {
                    "client_id": "hardware-1",
                    "parent_client_id": "component-1",
                    "kind": "hardware",
                    "name": "主控板",
                    "is_prototype": True,
                },
                {
                    "client_id": "machine-1",
                    "kind": "machine",
                    "name": "批量验证机",
                    "is_prototype": False,
                },
                {
                    "client_id": "structure-1",
                    "parent_client_id": "component-1",
                    "kind": "structure",
                    "name": "机箱",
                    "is_prototype": False,
                },
            ],
        },
        headers={"X-CSRF-Token": csrf},
    )
    assert generated.status_code == 200, generated.text
    codes = {node["name"]: node["code"] for node in generated.json()["nodes"]}
    assert codes == {
        "批量验证机": "GH9753-PLYZJ-00-G-1.00",
        "主控部组件": "GH9753-PLYZJ-01-00-G-1.00",
        "主控板": "GH9753-PLYZJ-01-10-Z-2.00",
        "处理器": "GH9753-PLYZJ-01-10-00-G-1.00",
        "机箱": "GH9753-PLYZJ-01-01-G-1.00",
    }


def test_invalid_component_tree_rolls_back_entire_batch(client: TestClient) -> None:
    csrf = login(client, "user", "component-tree-rollback")
    invalid = client.post(
        "/api/component-codes/projects/tree/generate",
        json={
            "project_code": "9754",
            "nodes": [
                {"client_id": "machine", "kind": "machine", "name": "回滚验证机"},
                {
                    "client_id": "software",
                    "parent_client_id": "machine",
                    "kind": "software",
                    "name": "错误层级软件",
                },
            ],
        },
        headers={"X-CSRF-Token": csrf},
    )
    assert invalid.status_code == 400
    assert client.get("/api/component-codes/projects/9754").status_code == 404


def test_pure_structure_project_starts_at_component_without_machine(client: TestClient) -> None:
    csrf = login(client, "user", "pure-structure")
    generated = client.post(
        "/api/component-codes/projects/tree/generate",
        json={
            "project_code": "2460",
            "product_type": "structure",
            "nodes": [
                {"client_id": "component", "kind": "component", "name": "结构部组件"},
                {
                    "client_id": "structure",
                    "parent_client_id": "component",
                    "kind": "structure",
                    "name": "机箱结构",
                },
            ],
        },
        headers={"X-CSRF-Token": csrf},
    )
    assert generated.status_code == 200, generated.text
    body = generated.json()
    assert body["product_type"] == "structure"
    codes = {node["kind"]: node["code"] for node in body["nodes"]}
    assert codes == {
        "component": "GH2460-01-00-G-1.00",
        "structure": "GH2460-01-01-G-1.00",
    }


def test_pure_hardware_project_starts_at_component_without_machine(client: TestClient) -> None:
    csrf = login(client, "user", "pure-hardware")
    generated = client.post(
        "/api/component-codes/projects/tree/generate",
        json={
            "project_code": "2461",
            "product_type": "hardware",
            "nodes": [
                {"client_id": "component", "kind": "component", "name": "板卡部组件"},
                {
                    "client_id": "hardware",
                    "parent_client_id": "component",
                    "kind": "hardware",
                    "name": "主控板",
                },
            ],
        },
        headers={"X-CSRF-Token": csrf},
    )
    assert generated.status_code == 200, generated.text
    codes = {node["kind"]: node["code"] for node in generated.json()["nodes"]}
    assert codes == {
        "component": "GH2461-01-00-G-1.00",
        "hardware": "GH2461-01-10-G-1.00",
    }


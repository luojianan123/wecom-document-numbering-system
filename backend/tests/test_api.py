from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.exc import IntegrityError

from app.db import SessionLocal
from app.models import (
    CodeClaim,
    CodeReservation,
    FileCode,
    Project,
    ProjectBatchItem,
    User,
)

from .conftest import login


def test_health_and_mock_auth(client: TestClient) -> None:
    health = client.get("/api/health")
    assert health.status_code == 200
    assert health.json()["abbreviationEntries"] >= 120

    start = client.get("/api/auth/wecom/qr/start")
    assert start.status_code == 200
    assert start.json() == {"mode": "mock", "authorization_url": None}

    assert client.get("/api/me").status_code == 401
    csrf = login(client)
    me = client.get("/api/me")
    assert me.status_code == 200
    assert me.json()["user"]["role"] == "user"
    assert me.json()["csrf_token"] == csrf


def test_complete_stage_one_business_flow(client: TestClient) -> None:
    admin_csrf = login(client, "admin")
    csv_content = (
        "文件名称\n"
        "控制模块技术要求\n"
        "控制模块正样件技术要求\n"
        "控制PCB加工要求\n"
    ).encode()
    initialized = client.post(
        "/api/admin/projects/init",
        data={"project_name": "测试项目", "project_code": "1234"},
        files={"file": ("files.csv", csv_content, "text/csv")},
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert initialized.status_code == 200, initialized.text
    batch = initialized.json()
    assert batch["success_count"] == 3
    assert batch["failure_count"] == 0
    assert batch["items"][0]["final_code"] == "GH1234-3KZ-010JY-1.00"
    assert batch["items"][1]["final_code"] == "GH1234-3KZ-010JY-Z-2.00"
    assert batch["items"][2]["final_code"] == "GH1234-7KZ-010PB-1.00"

    project_id = batch["project"]["id"]
    draft_codes = client.get(f"/api/projects/{project_id}/codes")
    assert draft_codes.status_code == 404

    confirmed = client.post(
        f"/api/admin/projects/{project_id}/confirm",
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert confirmed.status_code == 200
    assert confirmed.json()["status"] == "active"

    user_csrf = login(client, "user")
    projects = client.get("/api/projects")
    assert projects.status_code == 200
    assert projects.json()[0]["project_code"] == "1234"

    all_codes = client.get(f"/api/projects/{project_id}/codes")
    assert all_codes.status_code == 200
    assert len(all_codes.json()) == 3
    assert [item["standard_name"] for item in all_codes.json()] == sorted(
        item["standard_name"] for item in all_codes.json()
    )
    assert client.get("/api/projects/999999/codes").status_code == 404

    search = client.get(
        "/api/codes/search",
        params={"project_id": project_id, "name": "控制模块"},
    )
    assert search.status_code == 200
    assert len(search.json()) == 2

    claimed = client.post(
        f"/api/codes/{search.json()[0]['id']}/claim",
        headers={"X-CSRF-Token": user_csrf},
    )
    assert claimed.status_code == 200
    assert claimed.json()["file_code"]["final_code"].startswith("GH1234-")
    assert claimed.json()["claimant_name"] == "普通用户"
    assert claimed.json()["claimed_at"]

    with SessionLocal() as db:
        claim = db.scalar(
            select(CodeClaim).where(
                CodeClaim.file_code_id == search.json()[0]["id"]
            )
        )
        assert claim is not None
        assert claim.claimant_name == "普通用户"
        claim.user.name = "后来改名"
        db.commit()

    admin_csrf = login(client, "admin")
    detail = client.get(f"/api/admin/projects/{project_id}")
    assert detail.status_code == 200, detail.text
    claimed_item = next(
        item
        for item in detail.json()["items"]
        if item["file_code_id"] == search.json()[0]["id"]
    )
    assert claimed_item["claims"][0]["claimant_name"] == "普通用户"
    assert claimed_item["claims"][0]["claimed_at"]

    user_csrf = login(client, "user")
    generated = client.post(
        "/api/codes/generate",
        json={
            "project_id": project_id,
            "file_name": "通信模块使用说明书.docx",
        },
        headers={"X-CSRF-Token": user_csrf},
    )
    assert generated.status_code == 200, generated.text
    assert generated.json()["status"] == "generated"
    assert generated.json()["file_code"]["final_code"] == "GH1234-3TX-010SS-1.00"
    assert generated.json()["file_code"]["standard_name"] == "通信模块使用说明书"
    assert generated.json()["file_code"]["enabled"] is True


def test_user_can_generate_requested_name_after_partial_match(
    client: TestClient,
) -> None:
    admin_csrf = login(client, "admin")
    initialized = client.post(
        "/api/admin/projects/init",
        data={"project_name": "相近文件查询项目", "project_code": "1235"},
        files={
            "file": (
                "files.csv",
                (
                    "文件名称\n"
                    "机载任务智能处理机S5000C主控板原理图\n"
                ).encode(),
                "text/csv",
            )
        },
        headers={"X-CSRF-Token": admin_csrf},
    )
    project_id = initialized.json()["project"]["id"]
    client.post(
        f"/api/admin/projects/{project_id}/confirm",
        headers={"X-CSRF-Token": admin_csrf},
    ).raise_for_status()

    user_csrf = login(client, "user")
    partial = client.get(
        "/api/codes/search",
        params={"project_id": project_id, "name": "主控板原理图"},
    )
    assert partial.status_code == 200
    assert len(partial.json()) == 1
    assert (
        partial.json()[0]["standard_name"]
        == "机载任务智能处理机S5000C主控板原理图"
    )

    generated = client.post(
        "/api/codes/generate",
        json={
            "project_id": project_id,
            "file_name": "主控板原理图",
        },
        headers={"X-CSRF-Token": user_csrf},
    )
    assert generated.status_code == 200, generated.text
    assert generated.json()["status"] == "pending_review"
    assert generated.json()["review"]["proposed_standard_name"] == "主控板原理图"
    assert generated.json()["review"]["similar_names"][0]["standard_name"] == (
        "机载任务智能处理机S5000C主控板原理图"
    )


def test_unrelated_project_name_is_submitted_for_admin_review(
    client: TestClient,
) -> None:
    admin_csrf = login(client, "admin")
    initialized = client.post(
        "/api/admin/projects/init",
        data={
            "project_name": "机上实时任务管控与应急响应分系统",
            "project_code": "1238",
        },
        files={
            "file": (
                "files.csv",
                (
                    "文件名称\n"
                    "机载任务智能处理机主控板原理图\n"
                    "机上实时任务管控软件测试报告\n"
                ).encode(),
                "text/csv",
            )
        },
        headers={"X-CSRF-Token": admin_csrf},
    )
    project_id = initialized.json()["project"]["id"]
    client.post(
        f"/api/admin/projects/{project_id}/confirm",
        headers={"X-CSRF-Token": admin_csrf},
    ).raise_for_status()

    user_csrf = login(client, "user")
    file_name = "我昨晚吃了好吃的水果开发计划"
    search = client.get(
        "/api/codes/search",
        params={"project_id": project_id, "name": file_name},
    )
    assert search.status_code == 200
    assert search.json() == []

    submitted = client.post(
        "/api/codes/generate",
        json={"project_id": project_id, "file_name": file_name},
        headers={"X-CSRF-Token": user_csrf},
    )
    assert submitted.status_code == 200, submitted.text
    assert submitted.json()["status"] == "pending_review"
    assert "疑似非工程内容" in submitted.json()["message"]
    assert "明显生活化" in submitted.json()["review"]["issue_summary"]
    assert submitted.json()["file_code"] is None


@pytest.mark.parametrize(
    "file_name",
    [
        "李京平是猪",
        "李京平原理图",
    ],
)
def test_personal_judgement_or_person_name_document_requires_review(
    client: TestClient,
    file_name: str,
) -> None:
    admin_csrf = login(client, "admin")
    initialized = client.post(
        "/api/admin/projects/init",
        data={
            "project_name": "压缩存储单元千兆网",
            "project_code": "1240",
        },
        files={
            "file": (
                "files.csv",
                "文件名称\n压缩存储单元千兆网正样件方案设计报告\n".encode(),
                "text/csv",
            )
        },
        headers={"X-CSRF-Token": admin_csrf},
    )
    project_id = initialized.json()["project"]["id"]
    client.post(
        f"/api/admin/projects/{project_id}/confirm",
        headers={"X-CSRF-Token": admin_csrf},
    ).raise_for_status()

    user_csrf = login(client, "user")
    submitted = client.post(
        "/api/codes/generate",
        json={"project_id": project_id, "file_name": file_name},
        headers={"X-CSRF-Token": user_csrf},
    )
    assert submitted.status_code == 200, submitted.text
    assert submitted.json()["status"] == "pending_review"
    assert submitted.json()["file_code"] is None
    assert "疑似非工程内容" in submitted.json()["message"]


def test_valid_new_component_name_is_generated_without_admin_review(
    client: TestClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    admin_notifications: list[dict[str, object]] = []

    async def fake_admin_notification(**payload: object) -> None:
        admin_notifications.append(payload)

    monkeypatch.setattr(
        "app.api.codes.notify_admin_review_requested",
        fake_admin_notification,
    )
    admin_csrf = login(client, "admin")
    initialized = client.post(
        "/api/admin/projects/init",
        data={
            "project_name": "压缩存储单元千兆网",
            "project_code": "1239",
        },
        files={
            "file": (
                "files.csv",
                "文件名称\n压缩存储单元千兆网正样件方案设计报告\n".encode(),
                "text/csv",
            )
        },
        headers={"X-CSRF-Token": admin_csrf},
    )
    project_id = initialized.json()["project"]["id"]
    client.post(
        f"/api/admin/projects/{project_id}/confirm",
        headers={"X-CSRF-Token": admin_csrf},
    ).raise_for_status()

    user_csrf = login(client, "user")
    generated = client.post(
        "/api/codes/generate",
        json={"project_id": project_id, "file_name": "主控板原理图"},
        headers={"X-CSRF-Token": user_csrf},
    )
    assert generated.status_code == 200, generated.text
    assert generated.json()["status"] == "generated"
    assert generated.json()["file_code"]["standard_name"] == "主控板原理图"
    assert generated.json()["file_code"]["final_code"].startswith("GH1239-5ZK-")
    assert generated.json()["review"] is None
    assert admin_notifications == []


@pytest.mark.parametrize(
    "file_name,message",
    [
        ("123456", "纯数字"),
        ("哈哈哈哈", "重复字符"),
        ("https://example.com", "网址"),
        ("控制模块😀技术要求", "表情"),
        ("控制模块@技术要求", "特殊符号"),
        ("asdf", "明显无关"),
    ],
)
def test_user_file_name_validation_rejects_invalid_content(
    client: TestClient,
    file_name: str,
    message: str,
) -> None:
    admin_csrf = login(client, "admin")
    initialized = client.post(
        "/api/admin/projects/init",
        data={"project_name": "名称校验项目", "project_code": "1236"},
        files={
            "file": (
                "files.csv",
                "文件名称\n控制模块技术要求\n".encode(),
                "text/csv",
            )
        },
        headers={"X-CSRF-Token": admin_csrf},
    )
    project_id = initialized.json()["project"]["id"]
    client.post(
        f"/api/admin/projects/{project_id}/confirm",
        headers={"X-CSRF-Token": admin_csrf},
    ).raise_for_status()

    user_csrf = login(client, "user")
    response = client.post(
        "/api/codes/generate",
        json={"project_id": project_id, "file_name": file_name},
        headers={"X-CSRF-Token": user_csrf},
    )
    assert response.status_code == 400
    assert message in response.json()["detail"]


def test_admin_approves_similar_name_review_then_user_receives_code(
    client: TestClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    admin_notifications: list[dict[str, object]] = []
    user_notifications: list[dict[str, object]] = []

    async def fake_admin_notification(**payload: object) -> None:
        admin_notifications.append(payload)

    async def fake_user_notification(**payload: object) -> None:
        user_notifications.append(payload)

    monkeypatch.setattr(
        "app.api.codes.notify_admin_review_requested",
        fake_admin_notification,
    )
    monkeypatch.setattr(
        "app.api.admin.notify_review_approved",
        fake_user_notification,
    )
    admin_csrf = login(client, "admin")
    initialized = client.post(
        "/api/admin/projects/init",
        data={"project_name": "名称审核项目", "project_code": "1237"},
        files={
            "file": (
                "files.csv",
                "文件名称\n机载任务智能处理机S5000C主控板原理图\n".encode(),
                "text/csv",
            )
        },
        headers={"X-CSRF-Token": admin_csrf},
    )
    project_id = initialized.json()["project"]["id"]
    client.post(
        f"/api/admin/projects/{project_id}/confirm",
        headers={"X-CSRF-Token": admin_csrf},
    ).raise_for_status()

    user_csrf = login(client, "user")
    submitted = client.post(
        "/api/codes/generate",
        json={"project_id": project_id, "file_name": "主控板原理图"},
        headers={"X-CSRF-Token": user_csrf},
    )
    assert submitted.status_code == 200
    review_id = submitted.json()["review"]["id"]
    assert len(admin_notifications) == 1
    assert admin_notifications[0]["review_id"] == review_id
    assert admin_notifications[0]["requester_user_id"] == "user-001"

    submitted_again = client.post(
        "/api/codes/generate",
        json={"project_id": project_id, "file_name": "主控板原理图"},
        headers={"X-CSRF-Token": user_csrf},
    )
    assert submitted_again.status_code == 200
    assert submitted_again.json()["review"]["id"] == review_id
    assert len(admin_notifications) == 1

    admin_csrf = login(client, "admin")
    reviews = client.get("/api/admin/name-reviews")
    assert reviews.status_code == 200
    assert [item["id"] for item in reviews.json()] == [review_id]

    approved = client.post(
        f"/api/admin/name-reviews/{review_id}/approve",
        json={"file_name": "备用主控板原理图"},
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert approved.status_code == 200, approved.text
    assert approved.json()["status"] == "approved"
    assert approved.json()["reviewed_name"] == "备用主控板原理图"
    assert approved.json()["file_code"]["final_code"].startswith("GH1237-")
    assert len(user_notifications) == 1
    assert user_notifications[0]["recipient_user_id"] == "user-001"
    assert user_notifications[0]["reviewed_name"] == "备用主控板原理图"

    user_csrf = login(client, "user")
    mine = client.get("/api/name-reviews/mine")
    assert mine.status_code == 200
    assert mine.json()[0]["status"] == "approved"
    assert mine.json()[0]["file_code"]["standard_name"] == "备用主控板原理图"

    admin_csrf = login(client, "admin")
    deleted = client.delete(
        (
            f"/api/admin/projects/{project_id}/codes/"
            f"{approved.json()['file_code_id']}"
        ),
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert deleted.status_code == 204

    user_csrf = login(client, "user")
    after_delete = client.get("/api/name-reviews/mine")
    assert after_delete.status_code == 200
    assert after_delete.json()[0]["status"] == "rejected"
    assert after_delete.json()[0]["file_code"] is None


def test_special_numbering_project_requires_admin_manual_review(
    client: TestClient,
) -> None:
    admin_csrf = login(client, "admin")
    initialized = client.post(
        "/api/admin/projects/init",
        data={
            "project_name": "特殊编号项目",
            "project_code": "1241",
            "special_numbering": "true",
        },
        files={
            "file": (
                "files.csv",
                "文件名称\n控制模块技术要求\n".encode(),
                "text/csv",
            )
        },
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert initialized.status_code == 200, initialized.text
    project = initialized.json()["project"]
    assert project["special_numbering"] is True
    project_id = project["id"]

    client.post(
        f"/api/admin/projects/{project_id}/confirm",
        headers={"X-CSRF-Token": admin_csrf},
    ).raise_for_status()

    admin_generated = client.post(
        f"/api/admin/projects/{project_id}/codes",
        json={"file_name": "通信模块使用说明书"},
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert admin_generated.status_code == 200, admin_generated.text
    assert admin_generated.json()["success"] is True

    user_csrf = login(client, "user")
    existing = client.post(
        "/api/codes/generate",
        json={"project_id": project_id, "file_name": "控制模块技术要求"},
        headers={"X-CSRF-Token": user_csrf},
    )
    assert existing.status_code == 200
    assert existing.json()["status"] == "existing"

    requested = client.post(
        "/api/codes/generate",
        json={"project_id": project_id, "file_name": "显示模块使用说明书"},
        headers={"X-CSRF-Token": user_csrf},
    )
    assert requested.status_code == 200, requested.text
    assert requested.json()["status"] == "pending_review"
    assert "等待管理员人工编号" in requested.json()["message"]
    review_id = requested.json()["review"]["id"]
    assert requested.json()["review"]["project"]["special_numbering"] is True

    admin_csrf = login(client, "admin")
    missing_code = client.post(
        f"/api/admin/name-reviews/{review_id}/approve",
        json={"file_name": "显示模块使用说明书"},
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert missing_code.status_code == 400
    assert "填写完整编号" in missing_code.json()["detail"]

    approved = client.post(
        f"/api/admin/name-reviews/{review_id}/approve",
        json={
            "file_name": "显示模块使用说明书",
            "final_code": "SP-1241-DISPLAY-001",
        },
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert approved.status_code == 200, approved.text
    assert approved.json()["status"] == "approved"
    assert approved.json()["file_code"]["final_code"] == "SP-1241-DISPLAY-001"
    assert approved.json()["file_code"]["source"] == "user_review_manual"

    user_csrf = login(client, "user")
    mine = client.get("/api/name-reviews/mine")
    result = next(item for item in mine.json() if item["id"] == review_id)
    assert result["file_code"]["final_code"] == "SP-1241-DISPLAY-001"

    admin_csrf = login(client, "admin")
    toggled = client.post(
        f"/api/admin/projects/{project_id}/special-numbering",
        json={"special_numbering": False},
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert toggled.status_code == 200
    assert toggled.json()["special_numbering"] is False


def test_software_batch_uses_level_five_and_r_prefix(
    client: TestClient,
) -> None:
    admin_csrf = login(client, "admin")
    initialized = client.post(
        "/api/admin/projects/init",
        data={"project_name": "软件项目", "project_code": "4321"},
        files={
            "file": (
                "files.csv",
                "文件名称\n控制模块软件配置管理计划\n".encode(),
                "text/csv",
            )
        },
        headers={"X-CSRF-Token": admin_csrf},
    )

    assert initialized.status_code == 200, initialized.text
    item = initialized.json()["items"][0]
    assert item["success"] is True
    assert item["final_code"] == "R-GH4321-5KZ-010SCP-1.00"


def test_unknown_software_document_uses_fallback_abbreviation(
    client: TestClient,
) -> None:
    admin_csrf = login(client, "admin")
    initialized = client.post(
        "/api/admin/projects/init",
        data={"project_name": "备用简号项目", "project_code": "4322"},
        files={
            "file": (
                "files.csv",
                "文件名称\n压缩存储单元主控软件安全性计划\n".encode(),
                "text/csv",
            )
        },
        headers={"X-CSRF-Token": admin_csrf},
    )

    assert initialized.status_code == 200, initialized.text
    item = initialized.json()["items"][0]
    assert item["success"] is True
    assert item["final_code"].startswith("R-GH4322-5CC-010")
    assert item["final_code"].endswith("-1.00")


def test_admin_can_manually_add_code_when_ai_cannot_match(
    client: TestClient,
) -> None:
    admin_csrf = login(client, "admin")
    initialized = client.post(
        "/api/admin/projects/init",
        data={"project_name": "手工补码项目", "project_code": "8765"},
        files={
            "file": (
                "files.csv",
                "文件名称\n控制模块技术要求\n".encode(),
                "text/csv",
            )
        },
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert initialized.status_code == 200, initialized.text
    project_id = initialized.json()["project"]["id"]

    staged = client.post(
        f"/api/admin/projects/{project_id}/codes/manual",
        json={
            "file_name": "未收录控制文件",
            "final_code": "GH8765-3KZ-010X1-1.00",
        },
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert staged.status_code == 200, staged.text
    assert staged.json()["file_code_id"] is None
    assert staged.json()["final_code"] == "GH8765-3KZ-010X1-1.00"

    duplicate = client.post(
        f"/api/admin/projects/{project_id}/codes/manual",
        json={
            "file_name": "另一未收录文件",
            "final_code": "GH8765-3KZ-010X1-1.00",
        },
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert duplicate.status_code == 409

    confirmed = client.post(
        f"/api/admin/projects/{project_id}/confirm",
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert confirmed.status_code == 200

    staged_software = client.post(
        f"/api/admin/projects/{project_id}/codes/manual",
        json={
            "file_name": "未收录软件文件",
            "final_code": "R-GH8765-5KZ-010X2-1.00",
        },
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert staged_software.status_code == 200, staged_software.text
    assert staged_software.json()["file_code_id"] is None

    confirmed_again = client.post(
        f"/api/admin/projects/{project_id}/confirm",
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert confirmed_again.status_code == 200
    with SessionLocal() as db:
        record = db.scalar(
            select(FileCode).where(
                FileCode.final_code == "R-GH8765-5KZ-010X2-1.00"
            )
        )
        assert record is not None
        assert record.source == "admin_batch"

    review_report = client.post(
        f"/api/admin/projects/{project_id}/codes/manual",
        json={
            "file_name": "控制模块评审结论报告",
            "final_code": "P-GH8765-3KZ-010X3-1.00",
        },
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert review_report.status_code == 200, review_report.text
    assert (
        review_report.json()["final_code"]
        == "P-GH8765-3KZ-010X3-1.00"
    )
    assert review_report.json()["file_code_id"] is None


def test_admin_and_csrf_guards(client: TestClient) -> None:
    user_csrf = login(client, "user")
    forbidden = client.get("/api/admin/projects")
    assert forbidden.status_code == 403

    no_csrf = client.post("/api/auth/logout")
    assert no_csrf.status_code == 403

    logout = client.post(
        "/api/auth/logout",
        headers={"X-CSRF-Token": user_csrf},
    )
    assert logout.status_code == 204
    assert client.get("/api/me").status_code == 401


def test_invalid_qr_callback_returns_to_login_with_retry_message(
    client: TestClient,
) -> None:
    response = client.get(
        "/api/auth/wecom/qr/callback",
        params={"code": "expired-code", "state": "invalid-state"},
        follow_redirects=False,
    )

    assert response.status_code == 303
    assert response.headers["location"].startswith("http://localhost:5173/login?")
    assert "auth_error=" in response.headers["location"]


def test_admin_can_manually_fix_and_retry_failed_batch_item(
    client: TestClient,
) -> None:
    admin_csrf = login(client, "admin")
    initialized = client.post(
        "/api/admin/projects/init",
        data={"project_name": "手动修正项目", "project_code": "5678"},
        files={
            "file": (
                "files.csv",
                "文件名称\nA\n".encode(),
                "text/csv",
            )
        },
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert initialized.status_code == 200
    batch = initialized.json()
    assert batch["success_count"] == 0
    assert batch["failure_count"] == 1
    batch_item_id = batch["items"][0]["id"]

    pending_detail = client.get(
        f"/api/admin/projects/{batch['project']['id']}",
    )
    assert pending_detail.status_code == 200
    assert pending_detail.json()["project"]["status"] == "draft"
    assert pending_detail.json()["failure_count"] == 1

    retried = client.post(
        (
            f"/api/admin/projects/{batch['project']['id']}"
            f"/batch-items/{batch_item_id}/retry"
        ),
        json={"file_name": "控制模块技术要求"},
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert retried.status_code == 200
    assert retried.json()["success"] is True
    assert retried.json()["original_name"] == "A"
    assert retried.json()["standard_name"] == "控制模块技术要求"
    assert retried.json()["final_code"] == "GH5678-3KZ-010JY-1.00"

    confirmed = client.post(
        f"/api/admin/projects/{batch['project']['id']}/confirm",
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert confirmed.status_code == 200

    active_detail = client.get(
        f"/api/admin/projects/{batch['project']['id']}",
    )
    assert active_detail.status_code == 200
    assert active_detail.json()["project"]["status"] == "active"
    assert active_detail.json()["success_count"] == 1
    assert active_detail.json()["failure_count"] == 0


def test_admin_can_manually_number_failed_batch_item(
    client: TestClient,
) -> None:
    admin_csrf = login(client, "admin")
    initialized = client.post(
        "/api/admin/projects/init",
        data={"project_name": "人工编号项目", "project_code": "5680"},
        files={
            "file": (
                "files.csv",
                "文件名称\nA\n".encode(),
                "text/csv",
            )
        },
        headers={"X-CSRF-Token": admin_csrf},
    )
    batch = initialized.json()
    item = batch["items"][0]
    assert item["success"] is False

    manual = client.post(
        (
            f"/api/admin/projects/{batch['project']['id']}"
            f"/batch-items/{item['id']}/manual"
        ),
        json={
            "file_name": "人工控制文件",
            "final_code": "GH5680-3KZ-010RG-1.00",
        },
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert manual.status_code == 200, manual.text
    assert manual.json()["success"] is True
    assert manual.json()["standard_name"] == "人工控制文件"
    assert manual.json()["final_code"] == "GH5680-3KZ-010RG-1.00"

    confirmed = client.post(
        f"/api/admin/projects/{batch['project']['id']}/confirm",
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert confirmed.status_code == 200, confirmed.text

    detail = client.get(
        f"/api/admin/projects/{batch['project']['id']}",
    ).json()
    assert detail["success_count"] == 1
    assert detail["failure_count"] == 0
    assert detail["items"][0]["file_code_id"] is not None


def test_admin_can_edit_successful_pending_batch_item(
    client: TestClient,
) -> None:
    admin_csrf = login(client, "admin")
    initialized = client.post(
        "/api/admin/projects/init",
        data={"project_name": "行内修改项目", "project_code": "5681"},
        files={
            "file": (
                "files.csv",
                (
                    "文件名称\n"
                    "控制模块技术要求\n"
                    "通信模块使用说明书\n"
                ).encode(),
                "text/csv",
            )
        },
        headers={"X-CSRF-Token": admin_csrf},
    )
    batch = initialized.json()
    item = batch["items"][0]
    original_code = item["final_code"]
    occupied_code = batch["items"][1]["final_code"]
    updated_code = "GH5681-3KZ-010X1-1.00"

    conflict = client.post(
        (
            f"/api/admin/projects/{batch['project']['id']}"
            f"/batch-items/{item['id']}/manual"
        ),
        json={
            "file_name": "人工修正控制文件",
            "final_code": occupied_code,
        },
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert conflict.status_code == 409
    with SessionLocal() as db:
        assert db.get(CodeReservation, original_code) is not None

    modified = client.post(
        (
            f"/api/admin/projects/{batch['project']['id']}"
            f"/batch-items/{item['id']}/manual"
        ),
        json={
            "file_name": "人工修正控制文件",
            "final_code": updated_code,
        },
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert modified.status_code == 200, modified.text
    assert modified.json()["success"] is True
    assert modified.json()["standard_name"] == "人工修正控制文件"
    assert modified.json()["final_code"] == updated_code

    with SessionLocal() as db:
        assert db.get(CodeReservation, original_code) is None
        assert db.get(CodeReservation, updated_code) is not None

    confirmed = client.post(
        f"/api/admin/projects/{batch['project']['id']}/confirm",
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert confirmed.status_code == 200, confirmed.text
    with SessionLocal() as db:
        stored = db.scalar(
            select(FileCode).where(FileCode.final_code == updated_code)
        )
        assert stored is not None
        assert stored.standard_name == "人工修正控制文件"


def test_batch_successes_are_only_persisted_after_admin_confirmation(
    client: TestClient,
) -> None:
    admin_csrf = login(client, "admin")
    initialized = client.post(
        "/api/admin/projects/init",
        data={"project_name": "待确认项目", "project_code": "2468"},
        files={
            "file": (
                "files.csv",
                "文件名称\n控制模块技术要求\n通信模块使用说明书\n".encode(),
                "text/csv",
            )
        },
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert initialized.status_code == 200, initialized.text
    batch = initialized.json()
    project_id = batch["project"]["id"]
    assert batch["success_count"] == 2
    assert all(item["file_code_id"] is None for item in batch["items"])

    with SessionLocal() as db:
        assert db.scalar(select(func.count()).select_from(FileCode)) == 0
        staged = list(
            db.scalars(
                select(ProjectBatchItem).where(
                    ProjectBatchItem.project_id == project_id
                )
            )
        )
        assert len(staged) == 2
        assert all(item.preview_data for item in staged)
        assert all(item.file_code_id is None for item in staged)
        assert all(item.preview_final_code for item in staged)
        assert (
            db.scalar(select(func.count()).select_from(CodeReservation)) == 2
        )

        project = db.get(Project, project_id)
        assert project is not None
        other_project = Project(
            project_code="8642",
            project_name="全局查重测试项目",
            status="draft",
            created_by_id=project.created_by_id,
        )
        db.add(other_project)
        db.flush()
        db.add(
            CodeReservation(
                project_id=other_project.id,
                final_code=staged[0].preview_final_code,
            )
        )
        with pytest.raises(IntegrityError):
            db.commit()
        db.rollback()

    reopened = client.get(f"/api/admin/projects/{project_id}")
    assert reopened.status_code == 200
    assert reopened.json()["items"][0]["final_code"] == "GH2468-3KZ-010JY-1.00"
    assert reopened.json()["items"][0]["file_code_id"] is None

    confirmed = client.post(
        f"/api/admin/projects/{project_id}/confirm",
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert confirmed.status_code == 200, confirmed.text

    with SessionLocal() as db:
        codes = list(
            db.scalars(
                select(FileCode)
                .where(FileCode.project_id == project_id)
                .order_by(FileCode.id)
            )
        )
        assert len(codes) == 2
        assert all(code.enabled for code in codes)
        assert (
            db.scalar(select(func.count()).select_from(CodeReservation)) == 2
        )
        staged = list(
            db.scalars(
                select(ProjectBatchItem).where(
                    ProjectBatchItem.project_id == project_id
                )
            )
        )
        assert all(item.file_code_id is not None for item in staged)


def test_admin_can_manage_project_files_codes_and_delete_project(
    client: TestClient,
) -> None:
    admin_csrf = login(client, "admin")
    initialized = client.post(
        "/api/admin/projects/init",
        data={"project_name": "管理操作项目", "project_code": "1357"},
        files={
            "file": (
                "files.csv",
                "文件名称\n控制模块技术要求\n通信模块使用说明书\n".encode(),
                "text/csv",
            )
        },
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert initialized.status_code == 200
    batch = initialized.json()
    project_id = batch["project"]["id"]

    removed_staged = client.delete(
        f"/api/admin/projects/{project_id}/batch-items/{batch['items'][1]['id']}",
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert removed_staged.status_code == 204

    added_to_draft = client.post(
        f"/api/admin/projects/{project_id}/codes",
        json={"file_name": "通信模块使用说明书"},
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert added_to_draft.status_code == 200, added_to_draft.text
    assert added_to_draft.json()["file_code_id"] is None

    confirmed = client.post(
        f"/api/admin/projects/{project_id}/confirm",
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert confirmed.status_code == 200, confirmed.text
    detail = client.get(f"/api/admin/projects/{project_id}").json()
    first_code_id = detail["items"][0]["file_code_id"]
    assert first_code_id is not None

    user_csrf = login(client, "user")
    claimed = client.post(
        f"/api/codes/{first_code_id}/claim",
        headers={"X-CSRF-Token": user_csrf},
    )
    assert claimed.status_code == 200

    admin_csrf = login(client, "admin")
    deleted_code = client.delete(
        f"/api/admin/projects/{project_id}/codes/{first_code_id}",
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert deleted_code.status_code == 204
    with SessionLocal() as db:
        assert db.get(FileCode, first_code_id) is None
        assert (
            db.scalar(
                select(func.count())
                .select_from(CodeClaim)
                .where(CodeClaim.file_code_id == first_code_id)
            )
            == 0
        )
        assert (
            db.scalar(
                select(func.count())
                .select_from(ProjectBatchItem)
                .where(ProjectBatchItem.project_id == project_id)
            )
            == 1
        )
        assert (
            db.scalar(
                select(func.count())
                .select_from(CodeReservation)
                .where(CodeReservation.project_id == project_id)
            )
            == 1
        )

    added_to_active = client.post(
        f"/api/admin/projects/{project_id}/codes",
        json={"file_name": "控制PCB加工要求"},
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert added_to_active.status_code == 200, added_to_active.text
    assert added_to_active.json()["file_code_id"] is None

    confirmed_added = client.post(
        f"/api/admin/projects/{project_id}/confirm",
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert confirmed_added.status_code == 200
    with SessionLocal() as db:
        added_record = db.scalar(
            select(FileCode).where(
                FileCode.project_id == project_id,
                FileCode.standard_name == "控制PCB加工要求",
            )
        )
        assert added_record is not None

    no_csrf = client.delete(f"/api/admin/projects/{project_id}")
    assert no_csrf.status_code == 403
    user_csrf = login(client, "user")
    forbidden = client.delete(
        f"/api/admin/projects/{project_id}",
        headers={"X-CSRF-Token": user_csrf},
    )
    assert forbidden.status_code == 403

    admin_csrf = login(client, "admin")
    deleted_project = client.delete(
        f"/api/admin/projects/{project_id}",
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert deleted_project.status_code == 204
    assert client.get(f"/api/admin/projects/{project_id}").status_code == 404
    with SessionLocal() as db:
        assert db.get(Project, project_id) is None
        assert (
            db.scalar(
                select(func.count())
                .select_from(FileCode)
                .where(FileCode.project_id == project_id)
            )
            == 0
        )
        assert (
            db.scalar(
                select(func.count())
                .select_from(CodeReservation)
                .where(CodeReservation.project_id == project_id)
            )
            == 0
        )
        assert (
            db.scalar(
                select(func.count())
                .select_from(ProjectBatchItem)
                .where(ProjectBatchItem.project_id == project_id)
            )
            == 0
        )


def test_active_project_import_marks_duplicates_and_stages_new_files(
    client: TestClient,
) -> None:
    admin_csrf = login(client, "admin")
    initialized = client.post(
        "/api/admin/projects/init",
        data={"project_name": "已有项目批量新增", "project_code": "1360"},
        files={
            "file": (
                "files.csv",
                "文件名称\n控制模块技术要求\n".encode(),
                "text/csv",
            )
        },
        headers={"X-CSRF-Token": admin_csrf},
    )
    project_id = initialized.json()["project"]["id"]
    confirmed = client.post(
        f"/api/admin/projects/{project_id}/confirm",
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert confirmed.status_code == 200

    imported = client.post(
        f"/api/admin/projects/{project_id}/codes/import",
        files={
            "file": (
                "new-files.csv",
                (
                    "文件名称\n"
                    "控制模块技术要求\n"
                    "通信模块使用说明书\n"
                ).encode(),
                "text/csv",
            )
        },
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert imported.status_code == 200, imported.text
    detail = imported.json()
    duplicate = next(
        item
        for item in detail["items"]
        if item["original_name"] == "控制模块技术要求"
        and item["file_code_id"] is None
    )
    pending = next(
        item
        for item in detail["items"]
        if item["original_name"] == "通信模块使用说明书"
    )
    assert duplicate["success"] is False
    assert duplicate["error"].startswith("已重复：")
    assert pending["success"] is True
    assert pending["file_code_id"] is None

    with SessionLocal() as db:
        assert (
            db.scalar(
                select(func.count())
                .select_from(FileCode)
                .where(FileCode.project_id == project_id)
            )
            == 1
        )

    confirmed_new = client.post(
        f"/api/admin/projects/{project_id}/confirm",
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert confirmed_new.status_code == 200
    with SessionLocal() as db:
        assert (
            db.scalar(
                select(func.count())
                .select_from(FileCode)
                .where(FileCode.project_id == project_id)
            )
            == 2
        )


def test_admin_can_export_all_stored_project_codes_to_excel(
    client: TestClient,
) -> None:
    admin_csrf = login(client, "admin")
    initialized = client.post(
        "/api/admin/projects/init",
        data={"project_name": "编码导出项目", "project_code": "1362"},
        files={
            "file": (
                "files.csv",
                (
                    "文件名称\n"
                    "控制模块技术要求\n"
                    "通信模块使用说明书\n"
                ).encode(),
                "text/csv",
            )
        },
        headers={"X-CSRF-Token": admin_csrf},
    )
    project_id = initialized.json()["project"]["id"]

    blocked = client.get(f"/api/admin/projects/{project_id}/export")
    assert blocked.status_code == 409

    confirmed = client.post(
        f"/api/admin/projects/{project_id}/confirm",
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert confirmed.status_code == 200

    exported = client.get(f"/api/admin/projects/{project_id}/export")
    assert exported.status_code == 200, exported.text
    assert exported.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "filename*=UTF-8''" in exported.headers["content-disposition"]

    workbook = load_workbook(BytesIO(exported.content), read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    claim_rows = list(
        workbook["领取记录"].iter_rows(values_only=True)
    )
    workbook.close()

    assert rows[0] == ("文件名称", "文件编号")
    assert rows[1:] == [
        ("控制模块技术要求", "GH1362-3KZ-010JY-1.00"),
        ("通信模块使用说明书", "GH1362-3TX-010SS-1.00"),
    ]
    assert claim_rows == [("文件名称", "文件编号", "领取人", "领取时间")]

    imported = client.post(
        f"/api/admin/projects/{project_id}/codes/import",
        files={
            "file": (
                "added.csv",
                (
                    "文件名称\n"
                    "控制模块技术要求\n"
                    "显示模块使用说明书\n"
                ).encode(),
                "text/csv",
            )
        },
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert imported.status_code == 200
    blocked_again = client.get(f"/api/admin/projects/{project_id}/export")
    assert blocked_again.status_code == 409
    assert "全部处理后才能导出" in blocked_again.json()["detail"]


def test_admin_can_batch_delete_mixed_project_files(
    client: TestClient,
) -> None:
    admin_csrf = login(client, "admin")
    initialized = client.post(
        "/api/admin/projects/init",
        data={"project_name": "批量删除项目", "project_code": "1361"},
        files={
            "file": (
                "files.csv",
                (
                    "文件名称\n"
                    "控制模块技术要求\n"
                    "通信模块使用说明书\n"
                ).encode(),
                "text/csv",
            )
        },
        headers={"X-CSRF-Token": admin_csrf},
    )
    project_id = initialized.json()["project"]["id"]
    client.post(
        f"/api/admin/projects/{project_id}/confirm",
        headers={"X-CSRF-Token": admin_csrf},
    ).raise_for_status()

    user_csrf = login(client, "user")
    stored_codes = client.get(f"/api/projects/{project_id}/codes").json()
    claimed_code = stored_codes[0]
    claim = client.post(
        f"/api/codes/{claimed_code['id']}/claim",
        headers={"X-CSRF-Token": user_csrf},
    )
    assert claim.status_code == 200

    admin_csrf = login(client, "admin")
    imported = client.post(
        f"/api/admin/projects/{project_id}/codes/import",
        files={
            "file": (
                "added.csv",
                (
                    "文件名称\n"
                    "控制模块技术要求\n"
                    "显示模块使用说明书\n"
                ).encode(),
                "text/csv",
            )
        },
        headers={"X-CSRF-Token": admin_csrf},
    ).json()
    duplicate = next(
        item
        for item in imported["items"]
        if not item["success"] and item["error"].startswith("已重复：")
    )
    pending = next(
        item
        for item in imported["items"]
        if item["success"] and item["file_code_id"] is None
    )

    deleted = client.post(
        f"/api/admin/projects/{project_id}/files/batch-delete",
        json={
            "file_code_ids": [claimed_code["id"]],
            "batch_item_ids": [duplicate["id"], pending["id"]],
        },
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert deleted.status_code == 204, deleted.text

    with SessionLocal() as db:
        assert db.get(FileCode, claimed_code["id"]) is None
        assert (
            db.scalar(
                select(func.count())
                .select_from(CodeClaim)
                .where(CodeClaim.file_code_id == claimed_code["id"])
            )
            == 0
        )
        assert db.get(ProjectBatchItem, duplicate["id"]) is None
        assert db.get(ProjectBatchItem, pending["id"]) is None
        assert db.get(CodeReservation, pending["final_code"]) is None

        remaining_codes = list(
            db.scalars(
                select(FileCode).where(FileCode.project_id == project_id)
            )
        )
        assert len(remaining_codes) == 1


def test_initializing_project_cannot_be_confirmed_or_modified_but_can_be_deleted(
    client: TestClient,
) -> None:
    admin_csrf = login(client, "admin")
    with SessionLocal() as db:
        admin = db.scalar(select(User).where(User.role == "admin"))
        assert admin is not None
        project = Project(
            project_code="9090",
            project_name="生成中的项目",
            status="initializing",
            created_by_id=admin.id,
        )
        db.add(project)
        db.commit()
        db.refresh(project)
        project_id = project.id

    confirmed = client.post(
        f"/api/admin/projects/{project_id}/confirm",
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert confirmed.status_code == 409

    added = client.post(
        f"/api/admin/projects/{project_id}/codes",
        json={"file_name": "控制模块技术要求"},
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert added.status_code == 409

    deleted = client.delete(
        f"/api/admin/projects/{project_id}",
        headers={"X-CSRF-Token": admin_csrf},
    )
    assert deleted.status_code == 204
    with SessionLocal() as db:
        assert db.get(Project, project_id) is None

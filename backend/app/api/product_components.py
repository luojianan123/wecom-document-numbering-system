from io import BytesIO
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from sqlalchemy import delete, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from ..db import get_db
from ..models import ComponentClaim, ComponentNode, ComponentProject, User
from ..schemas import (
    ComponentBulkDeleteIn,
    ComponentClaimOut,
    ComponentDraftNodeIn,
    ComponentMachineCreateIn,
    ComponentNodeCreateIn,
    ComponentNodeOut,
    ComponentNodeUpdateIn,
    ComponentProjectCreateIn,
    ComponentProjectOut,
    ComponentProjectSummaryOut,
    ComponentTreeGenerateIn,
)
from ..security import CurrentSession, require_admin, require_csrf, require_user
from ..services.product_components import (
    ComponentNumberingError,
    build_child_code,
    build_component_root_code,
    build_machine_code,
    kind_label,
    rebuild_code_suffix,
    sequence_for_index,
    stage_code,
    validate_node_code,
)

router = APIRouter(prefix="/api/component-codes", tags=["产品组件编码"])


def _node_out(node: ComponentNode) -> ComponentNodeOut:
    return ComponentNodeOut(
        id=node.id,
        component_project_id=node.component_project_id,
        parent_id=node.parent_id,
        kind=node.kind,
        name=node.name,
        code=node.code,
        stage=node.stage,
        sequence=node.sequence,
        created_by_name=node.created_by.name if node.created_by else "未知用户",
        claims=[
            ComponentClaimOut(
                id=claim.id,
                claimant_name=claim.claimant_name,
                claimed_at=claim.claimed_at,
            )
            for claim in sorted(node.claims, key=lambda item: item.claimed_at, reverse=True)
        ],
    )


def _project_out(db: Session, project: ComponentProject) -> ComponentProjectOut:
    nodes = list(
        db.scalars(
            select(ComponentNode)
            .where(ComponentNode.component_project_id == project.id)
            .order_by(ComponentNode.id)
        )
    )
    return ComponentProjectOut(
        id=project.id,
        project_code=project.project_code,
        product_type=project.product_type,
        status=project.status,
        created_at=project.created_at,
        created_by_name=project.created_by.name if project.created_by else "未知用户",
        nodes=[_node_out(node) for node in nodes],
    )


def _get_project(db: Session, project_id: int) -> ComponentProject:
    project = db.get(ComponentProject, project_id)
    if not project:
        raise HTTPException(status_code=404, detail="产品组件项目不存在")
    return project


def _get_node(db: Session, project_id: int, node_id: int) -> ComponentNode:
    node = db.scalar(
        select(ComponentNode).where(
            ComponentNode.id == node_id,
            ComponentNode.component_project_id == project_id,
        )
    )
    if not node:
        raise HTTPException(status_code=404, detail="产品组件不存在")
    return node


def _next_sequence(db: Session, parent_id: int, kind: str) -> int:
    used = set(
        db.scalars(
            select(ComponentNode.sequence).where(
                ComponentNode.parent_id == parent_id,
            )
        )
    )
    index = 0
    while True:
        sequence = sequence_for_index(kind, index)
        if sequence > 99:
            raise ComponentNumberingError("该层级两位序列号已用完")
        if sequence not in used:
            return sequence
        index += 1


def _create_draft_node(
    db: Session,
    project: ComponentProject,
    payload: ComponentDraftNodeIn,
    parent: ComponentNode | None,
    user: User,
) -> ComponentNode:
    stage = stage_code(payload.stage, payload.is_prototype)
    name = payload.name.strip()
    if payload.kind == "machine":
        if project.product_type != "machine":
            raise ComponentNumberingError("纯结构件或纯板卡项目不能增加整机")
        if parent is not None:
            raise ComponentNumberingError("整机不能设置上级")
        if db.scalar(
            select(ComponentNode).where(
                ComponentNode.component_project_id == project.id,
                ComponentNode.parent_id.is_(None),
                ComponentNode.name == name,
            )
        ):
            raise ComponentNumberingError(f"整机名称“{name}”已存在")
        code = build_machine_code(project.project_code, name, stage)
        sequence = 0
    elif (
        payload.kind == "component"
        and parent is None
        and project.product_type in {"structure", "hardware"}
    ):
        if db.scalar(
            select(ComponentNode).where(
                ComponentNode.component_project_id == project.id,
                ComponentNode.parent_id.is_(None),
            )
        ):
            raise ComponentNumberingError("该项目的部组件根节点已存在")
        code = build_component_root_code(project.project_code, stage)
        sequence = 1
    else:
        if parent is None:
            raise ComponentNumberingError(f"{kind_label(payload.kind)}必须设置上级")
        sequence = _next_sequence(db, parent.id, payload.kind)
        code = build_child_code(parent, payload.kind, sequence, stage)
    node = ComponentNode(
        component_project_id=project.id,
        parent_id=parent.id if parent else None,
        kind=payload.kind,
        name=name,
        code=code,
        stage=stage,
        sequence=sequence,
        created_by_id=user.id,
    )
    db.add(node)
    db.flush()
    return node


def _descendant_ids(db: Session, project_id: int, roots: set[int]) -> set[int]:
    all_nodes = list(
        db.scalars(
            select(ComponentNode).where(ComponentNode.component_project_id == project_id)
        )
    )
    selected = set(roots)
    changed = True
    while changed:
        changed = False
        for node in all_nodes:
            if node.parent_id in selected and node.id not in selected:
                selected.add(node.id)
                changed = True
    return selected


def _renumber_descendants(db: Session, parent: ComponentNode) -> list[ComponentNode]:
    """Return descendants after assigning contiguous sibling sequences.

    Codes are assigned in a second pass by ``_renumber_tree``.  Keeping the
    sequence pass separate prevents an intermediate code from colliding with
    another node's old code while SQLAlchemy flushes the transaction.
    """
    children = list(
        db.scalars(
            select(ComponentNode)
            .where(ComponentNode.parent_id == parent.id)
            .order_by(ComponentNode.sequence, ComponentNode.id)
        )
    )
    by_kind: dict[str, list[ComponentNode]] = {}
    for child in children:
        by_kind.setdefault(child.kind, []).append(child)
    changed: list[ComponentNode] = []
    for kind, siblings in by_kind.items():
        for offset, child in enumerate(siblings):
            child.sequence = sequence_for_index(kind, offset)
            changed.append(child)
            changed.extend(_renumber_descendants(db, child))
    return changed


def _renumber_tree(db: Session, machines: list[ComponentNode]) -> None:
    changed: list[ComponentNode] = []
    for machine in machines:
        changed.extend(_renumber_descendants(db, machine))
    if not machines:
        return
    # Use temporary unique codes first.  This makes a swap/reorder safe even
    # when the database enforces uniqueness on the code column.
    for index, node in enumerate(changed):
        node.code = f"__renumbering_{node.id}_{index}"
    db.flush()
    for machine in machines:
        machine.code = rebuild_code_suffix(machine)
    for node in changed:
        node.code = build_child_code(node.parent, node.kind, node.sequence, node.stage)


@router.get(
    "/projects/admin/list",
    response_model=list[ComponentProjectSummaryOut],
)
def list_admin_projects(
    _: User = Depends(require_admin),
    db: Session = Depends(get_db),
) -> list[ComponentProjectSummaryOut]:
    projects = list(
        db.scalars(select(ComponentProject).order_by(ComponentProject.created_at.desc()))
    )
    result: list[ComponentProjectSummaryOut] = []
    for project in projects:
        nodes = list(
            db.scalars(
                select(ComponentNode).where(
                    ComponentNode.component_project_id == project.id
                )
            )
        )
        result.append(
            ComponentProjectSummaryOut(
                id=project.id,
                project_code=project.project_code,
                product_type=project.product_type,
                status=project.status,
                created_at=project.created_at,
                created_by_name=project.created_by.name if project.created_by else "未知用户",
                machine_count=sum(node.kind == "machine" for node in nodes),
                node_count=len(nodes),
                claim_count=sum(len(node.claims) for node in nodes),
            )
        )
    return result


@router.get("/projects/{project_code}", response_model=ComponentProjectOut)
def get_project_by_code(
    project_code: str,
    _: User = Depends(require_user),
    db: Session = Depends(get_db),
) -> ComponentProjectOut:
    if len(project_code) != 4 or not project_code.isdigit():
        raise HTTPException(status_code=400, detail="项目号必须为4位数字")
    project = db.scalar(
        select(ComponentProject).where(ComponentProject.project_code == project_code)
    )
    if not project:
        raise HTTPException(status_code=404, detail="该项目号尚未建立产品组件编码")
    return _project_out(db, project)


@router.post("/projects", response_model=ComponentProjectOut)
def create_project(
    payload: ComponentProjectCreateIn,
    user: User = Depends(require_user),
    _: CurrentSession = Depends(require_csrf),
    db: Session = Depends(get_db),
) -> ComponentProjectOut:
    if db.scalar(
        select(ComponentProject).where(ComponentProject.project_code == payload.project_code)
    ):
        raise HTTPException(status_code=409, detail="该项目号已存在产品组件编码")
    project = ComponentProject(
        project_code=payload.project_code,
        product_type=payload.product_type,
        status="active",
        created_by_id=user.id,
    )
    db.add(project)
    db.flush()
    if not payload.machine_name:
        raise HTTPException(status_code=400, detail="创建项目时必须填写产品组成名称")
    stage = stage_code(payload.stage, payload.is_prototype)
    root_kind = "machine" if payload.product_type == "machine" else "component"
    root_code = (
        build_machine_code(payload.project_code, payload.machine_name, stage)
        if root_kind == "machine"
        else build_component_root_code(payload.project_code, stage)
    )
    root = ComponentNode(
        component_project_id=project.id,
        parent_id=None,
        kind=root_kind,
        name=payload.machine_name.strip(),
        code=root_code,
        stage=stage,
        sequence=0 if root_kind == "machine" else 1,
        created_by_id=user.id,
    )
    db.add(root)
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail="项目或整机编号已存在") from exc
    db.refresh(project)
    return _project_out(db, project)


@router.post("/projects/tree/generate", response_model=ComponentProjectOut)
def generate_component_tree(
    payload: ComponentTreeGenerateIn,
    user: User = Depends(require_user),
    _: CurrentSession = Depends(require_csrf),
    db: Session = Depends(get_db),
) -> ComponentProjectOut:
    project = db.scalar(
        select(ComponentProject).where(
            ComponentProject.project_code == payload.project_code
        )
    )
    if not project:
        project = ComponentProject(
            project_code=payload.project_code,
            product_type=payload.product_type,
            status="active",
            created_by_id=user.id,
        )
        db.add(project)
        db.flush()
    elif project.product_type != payload.product_type:
        raise HTTPException(status_code=409, detail="该项目的产品组成类型与当前选择不一致")

    client_ids = [item.client_id for item in payload.nodes]
    if len(client_ids) != len(set(client_ids)):
        raise HTTPException(status_code=400, detail="草稿节点标识不能重复")

    pending = list(payload.nodes)
    created: dict[str, ComponentNode] = {}
    try:
        while pending:
            progressed = False
            for item in pending.copy():
                if item.parent_id is not None and item.parent_client_id is not None:
                    raise ComponentNumberingError("一个节点只能选择一个上级")
                parent: ComponentNode | None = None
                if item.parent_id is not None:
                    parent = _get_node(db, project.id, item.parent_id)
                elif item.parent_client_id is not None:
                    parent = created.get(item.parent_client_id)
                    if parent is None:
                        continue
                created[item.client_id] = _create_draft_node(
                    db, project, item, parent, user
                )
                pending.remove(item)
                progressed = True
            if not progressed:
                raise ComponentNumberingError("草稿中存在无效或循环的层级关系")
        db.commit()
    except (ComponentNumberingError, IntegrityError) as exc:
        db.rollback()
        detail = (
            str(exc)
            if isinstance(exc, ComponentNumberingError)
            else "生成的产品组件编号已存在"
        )
        raise HTTPException(status_code=400, detail=detail) from exc
    db.refresh(project)
    return _project_out(db, project)


@router.post("/projects/{project_id}/machines", response_model=ComponentNodeOut)
def add_machine(
    project_id: int,
    payload: ComponentMachineCreateIn,
    user: User = Depends(require_user),
    _: CurrentSession = Depends(require_csrf),
    db: Session = Depends(get_db),
) -> ComponentNodeOut:
    project = _get_project(db, project_id)
    if project.product_type != "machine":
        raise HTTPException(status_code=400, detail="纯结构件或纯板卡项目不能增加整机")
    if db.scalar(
        select(ComponentNode).where(
            ComponentNode.component_project_id == project.id,
            ComponentNode.parent_id.is_(None),
            ComponentNode.name == payload.name.strip(),
        )
    ):
        raise HTTPException(status_code=409, detail="该项目下整机名称不能重复")
    stage = stage_code(payload.stage, payload.is_prototype)
    node = ComponentNode(
        component_project_id=project.id,
        parent_id=None,
        kind="machine",
        name=payload.name.strip(),
        code=build_machine_code(project.project_code, payload.name, stage),
        stage=stage,
        sequence=0,
        created_by_id=user.id,
    )
    db.add(node)
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail="整机名称生成的编号已存在") from exc
    db.refresh(node)
    return _node_out(node)


@router.post("/projects/{project_id}/nodes", response_model=ComponentNodeOut)
def add_node(
    project_id: int,
    payload: ComponentNodeCreateIn,
    user: User = Depends(require_user),
    _: CurrentSession = Depends(require_csrf),
    db: Session = Depends(get_db),
) -> ComponentNodeOut:
    _get_project(db, project_id)
    parent = _get_node(db, project_id, payload.parent_id)
    try:
        sequence = _next_sequence(db, parent.id, payload.kind)
        stage = stage_code(payload.stage, payload.is_prototype)
        code = build_child_code(parent, payload.kind, sequence, stage)
    except ComponentNumberingError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    node = ComponentNode(
        component_project_id=project_id,
        parent_id=parent.id,
        kind=payload.kind,
        name=payload.name.strip(),
        code=code,
        stage=stage,
        sequence=sequence,
        created_by_id=user.id,
    )
    db.add(node)
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail="生成的产品组件编号已存在") from exc
    db.refresh(node)
    return _node_out(node)


@router.post("/projects/{project_id}/nodes/{node_id}", response_model=ComponentNodeOut)
def update_node(
    project_id: int,
    node_id: int,
    payload: ComponentNodeUpdateIn,
    _: User = Depends(require_user),
    __: CurrentSession = Depends(require_csrf),
    db: Session = Depends(get_db),
) -> ComponentNodeOut:
    node = _get_node(db, project_id, node_id)
    project = _get_project(db, project_id)
    try:
        code, stage, sequence = validate_node_code(
            project.project_code, node, payload.code, node.parent
        )
    except ComponentNumberingError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    conflict = db.scalar(
        select(ComponentNode).where(
            ComponentNode.component_project_id == project_id,
            ComponentNode.code == code,
            ComponentNode.id != node.id,
        )
    )
    if conflict:
        raise HTTPException(status_code=409, detail="产品组件编号已存在")
    node.name = payload.name.strip()
    node.code = code
    node.stage = stage
    node.sequence = sequence
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail="产品组件编号已存在") from exc
    db.refresh(node)
    return _node_out(node)


@router.post("/projects/{project_id}/bulk-delete", status_code=204)
def bulk_delete_nodes(
    project_id: int,
    payload: ComponentBulkDeleteIn,
    _: User = Depends(require_user),
    __: CurrentSession = Depends(require_csrf),
    db: Session = Depends(get_db),
) -> None:
    _get_project(db, project_id)
    ids = _descendant_ids(db, project_id, set(payload.node_ids))
    if not ids:
        return
    db.execute(delete(ComponentClaim).where(ComponentClaim.component_node_id.in_(ids)))
    nodes = list(db.scalars(select(ComponentNode).where(ComponentNode.id.in_(ids))))
    for node in sorted(nodes, key=lambda item: item.id, reverse=True):
        db.delete(node)
    db.commit()


@router.post("/projects/{project_id}/renumber", response_model=ComponentProjectOut)
def renumber_project(
    project_id: int,
    _: User = Depends(require_user),
    __: CurrentSession = Depends(require_csrf),
    db: Session = Depends(get_db),
) -> ComponentProjectOut:
    project = _get_project(db, project_id)
    roots = list(
        db.scalars(
            select(ComponentNode)
            .where(
                ComponentNode.component_project_id == project.id,
                ComponentNode.parent_id.is_(None),
            )
            .order_by(ComponentNode.id)
        )
    )
    try:
        _renumber_tree(db, roots)
        db.commit()
    except (ComponentNumberingError, IntegrityError) as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail="重新生成产品组成编号失败") from exc
    return _project_out(db, project)


@router.post("/nodes/{node_id}/claim", response_model=ComponentClaimOut)
def claim_node(
    node_id: int,
    user: User = Depends(require_user),
    _: CurrentSession = Depends(require_csrf),
    db: Session = Depends(get_db),
) -> ComponentClaimOut:
    node = db.get(ComponentNode, node_id)
    if not node:
        raise HTTPException(status_code=404, detail="产品组件不存在")
    claim = ComponentClaim(
        component_node_id=node.id,
        user_id=user.id,
        claimant_name=user.name,
    )
    db.add(claim)
    db.commit()
    db.refresh(claim)
    return ComponentClaimOut(
        id=claim.id,
        claimant_name=claim.claimant_name,
        claimed_at=claim.claimed_at,
    )


@router.get("/projects/{project_id}/export")
def export_project(
    project_id: int,
    _: User = Depends(require_user),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    project = _get_project(db, project_id)
    nodes = list(
        db.scalars(
            select(ComponentNode)
            .where(ComponentNode.component_project_id == project.id)
            .order_by(ComponentNode.id)
        )
    )
    workbook = Workbook()
    workbook.remove(workbook.active)
    roots = [node for node in nodes if node.parent_id is None]
    if not roots:
        raise HTTPException(status_code=400, detail="该项目没有可导出的产品组成")
    headers = (
        ["整机", "部组件", "结构/硬件/软件/其他", "零件"]
        if project.product_type == "machine"
        else ["部组件", "结构/硬件", "零件"]
    )
    column_count = len(headers) * 3
    used_titles: set[str] = set()
    for root_index, root in enumerate(roots, start=1):
        # Excel 工作表名称最多 31 个字符，且不能包含这些特殊字符。
        title = "".join("_" if char in r"[]:*?/\\" else char for char in root.name).strip()
        title = title[:31] or f"产品组成{root_index}"
        base_title = title
        suffix = 2
        while title in used_titles:
            tail = f"_{suffix}"
            title = f"{base_title[:31 - len(tail)]}{tail}"
            suffix += 1
        used_titles.add(title)
        sheet = workbook.create_sheet(title)
        sheet.append([f"项目号：{project.project_code}"])
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)
        heading_row: list[str] = []
        for heading in headers:
            heading_row.extend([heading, "", ""])
        sheet.append(heading_row)
        for index in range(len(headers)):
            start = index * 3 + 1
            sheet.merge_cells(start_row=2, start_column=start, end_row=2, end_column=start + 2)
        sheet.append(["名称", "编号", "阶段"] * len(headers))

        children: dict[int, list[ComponentNode]] = {}
        for node in nodes:
            if node.parent_id is not None:
                children.setdefault(node.parent_id, []).append(node)
        for child_nodes in children.values():
            child_nodes.sort(key=lambda item: (item.sequence, item.id))

        def append_branch(
            path: list[ComponentNode],
            current_sheet=sheet,
            current_children=children,
        ) -> None:
            row: list[str] = []
            for node in path:
                row.extend([
                    node.name,
                    node.code,
                    "正样件" if node.stage == "Z" else "其他",
                ])
            row.extend([""] * (column_count - len(row)))
            current_sheet.append(row)
            for child in current_children.get(path[-1].id, []):
                append_branch(path + [child])

        append_branch([root])
        for column in range(1, column_count + 1):
            sheet.column_dimensions[chr(64 + column)].width = 30 if column % 3 != 3 else 12
        sheet.freeze_panes = "A4"
    content = BytesIO()
    workbook.save(content)
    content.seek(0)
    filename = f"GH{project.project_code}-产品组件编码.xlsx"
    return StreamingResponse(
        content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"
        },
    )

import csv
from io import BytesIO, StringIO
from pathlib import PurePath

from openpyxl import load_workbook


class UploadError(ValueError):
    pass


MAX_ROWS = 2_000
MAX_FILE_BYTES = 10 * 1024 * 1024


def _deduplicate(values: list[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        value = value.strip()
        if value and value not in seen:
            seen.add(value)
            result.append(value)
    if not result:
        raise UploadError("清单中没有有效的文件名称")
    if len(result) > MAX_ROWS:
        raise UploadError(f"单次最多处理 {MAX_ROWS} 个文件名称")
    return result


def parse_file_names(filename: str, content: bytes) -> list[str]:
    if len(content) > MAX_FILE_BYTES:
        raise UploadError("清单文件不能超过 10 MB")
    suffix = PurePath(filename).suffix.lower()
    if suffix == ".xlsx":
        return _parse_xlsx(content)
    if suffix == ".csv":
        return _parse_csv(content)
    raise UploadError("仅支持 XLSX 或 CSV 清单")


def _parse_xlsx(content: bytes) -> list[str]:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    except Exception as exc:
        raise UploadError(f"无法读取 XLSX 清单：{exc}") from exc
    finally:
        if "workbook" in locals():
            workbook.close()

    if not rows:
        raise UploadError("XLSX 清单为空")
    header = [str(value).strip() if value is not None else "" for value in rows[0]]
    try:
        name_index = header.index("文件名称")
    except ValueError as exc:
        raise UploadError("XLSX 第一行必须包含“文件名称”列") from exc
    values = [
        str(row[name_index]).strip()
        for row in rows[1:]
        if name_index < len(row) and row[name_index] is not None
    ]
    return _deduplicate(values)


def _parse_csv(content: bytes) -> list[str]:
    text = ""
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if not text:
        raise UploadError("CSV 编码无法识别，请使用 UTF-8 或 GB18030")
    reader = csv.DictReader(StringIO(text))
    if not reader.fieldnames or "文件名称" not in reader.fieldnames:
        raise UploadError("CSV 第一行必须包含“文件名称”列")
    return _deduplicate([row.get("文件名称", "") for row in reader])

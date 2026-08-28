import re
import unicodedata
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..config import get_settings
from .document_rules import abbreviation_match_names


class AbbreviationError(ValueError):
    pass


@dataclass(frozen=True)
class AbbreviationMatch:
    alias: str
    code: str
    is_software: bool = False


FIXED_SUFFIX_ABBREVIATIONS: tuple[tuple[str, str], ...] = tuple(
    sorted(
        (
            ("技术状态管理计划", "CP"),
            ("调试作业指导书", "TS"),
            ("质量保证计划", "SQA"),
            ("质量保证大纲", "SQA"),
            ("配置管理计划", "SCP"),
            ("调试指导书", "TS"),
            ("作业指导书", "TS"),
            ("质保计划", "SQA"),
            ("质保大纲", "SQA"),
            ("开发计划", "SDP"),
            ("调试记录", "TJ"),
        ),
        key=lambda item: len(item[0]),
        reverse=True,
    )
)


def normalize_for_match(value: str) -> str:
    value = unicodedata.normalize("NFKC", value).casefold()
    return re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "", value)


class AbbreviationRegistry:
    def __init__(self, source: Path):
        if not source.exists():
            raise FileNotFoundError(f"文件简号规则不存在：{source}")
        self.source = source
        self._aliases: dict[str, AbbreviationMatch] = {}
        self._load()

    @property
    def entry_count(self) -> int:
        return len(self._aliases)

    def _add(self, alias: str, code: str, is_software: bool = False) -> None:
        alias = alias.strip()
        code = code.strip().upper()
        key = normalize_for_match(alias)
        if not key or not code or code == "-":
            return
        existing = self._aliases.get(key)
        if existing and (existing.code != code or existing.is_software != is_software):
            raise AbbreviationError(f"文件简号规则冲突：{alias} 存在不一致的简号或软件分类")
        self._aliases[key] = AbbreviationMatch(
            alias=alias,
            code=code,
            is_software=is_software,
        )

    @staticmethod
    def _split_inline(value: str) -> list[str]:
        parts = [item.strip() for item in re.split(r"[/／、]+", value) if item.strip()]
        if len(parts) < 2:
            return parts

        document_suffixes = (
            "作业指导书",
            "工艺要求",
            "技术要求",
            "要求",
            "汇总表",
            "配套表",
            "说明书",
            "申请报告",
            "设计报告",
            "测试计划",
            "测试说明",
            "测试报告",
            "试验报告",
            "验收报告",
            "细则",
            "报告",
            "大纲",
            "计划",
            "文件",
            "程序",
            "代码",
            "图",
            "表",
            "单",
            "书",
            "包",
        )
        shared_suffix = next(
            (suffix for suffix in document_suffixes if parts[-1].endswith(suffix)),
            "",
        )
        if not shared_suffix:
            # 原文会单独保留；无法判断共享后缀时不生成过短的模糊别名。
            return []

        expanded: list[str] = []
        for part in parts:
            has_own_suffix = any(part.endswith(suffix) for suffix in document_suffixes)
            expanded.append(part if has_own_suffix else part + shared_suffix)
        return expanded

    @staticmethod
    def _header_columns(sheet: Worksheet) -> tuple[int, int] | None:
        headers = {
            str(cell.value).strip(): cell.column
            for cell in sheet[1]
            if cell.value is not None and str(cell.value).strip()
        }
        name_column = headers.get("文件名称")
        code_column = headers.get("文件简号")
        if name_column is None or code_column is None:
            return None
        return name_column, code_column

    @staticmethod
    def _merged_cell_value(
        sheet: Worksheet,
        row: int,
        column: int,
    ) -> tuple[object | None, tuple[int, int, int, int] | None]:
        value = sheet.cell(row=row, column=column).value
        for merged_range in sheet.merged_cells.ranges:
            if (
                merged_range.min_row <= row <= merged_range.max_row
                and merged_range.min_col <= column <= merged_range.max_col
            ):
                value = sheet.cell(
                    row=merged_range.min_row,
                    column=merged_range.min_col,
                ).value
                return value, (
                    merged_range.min_row,
                    merged_range.max_row,
                    merged_range.min_col,
                    merged_range.max_col,
                )
        return value, None

    def _load(self) -> None:
        if self.source.suffix.lower() != ".xlsx":
            raise AbbreviationError("文件简号规则必须使用 文件简号.xlsx")

        workbook = load_workbook(self.source, data_only=True, read_only=False)
        try:
            selected: tuple[Worksheet, int, int] | None = None
            for sheet in workbook.worksheets:
                columns = self._header_columns(sheet)
                if columns:
                    selected = (sheet, *columns)
                    break
            if selected is None:
                raise AbbreviationError("文件简号.xlsx 中未找到“文件名称”和“文件简号”列")

            sheet, name_column, code_column = selected
            grouped_rows: dict[
                tuple[object, ...],
                tuple[str, bool, list[str]],
            ] = {}
            software_section = False
            software_marker_found = False
            for row in range(2, sheet.max_row + 1):
                raw_name = sheet.cell(row=row, column=name_column).value
                raw_code, merged_range = self._merged_cell_value(
                    sheet,
                    row,
                    code_column,
                )
                name = "" if raw_name is None else str(raw_name).strip()
                if normalize_for_match(name) == "以下为软件":
                    software_section = True
                    software_marker_found = True
                    continue
                code = "" if raw_code is None else str(raw_code).strip()
                if not name or not code or code == "-":
                    continue

                group_key: tuple[object, ...] = (
                    ("merged", *merged_range) if merged_range is not None else ("row", row)
                )
                if group_key not in grouped_rows:
                    grouped_rows[group_key] = (code, software_section, [])
                grouped_rows[group_key][2].append(name)

            if not software_marker_found:
                raise AbbreviationError("文件简号.xlsx 中未找到软件分界行“以下为软件：”")

            shared_suffixes = {"工艺要求", "技术要求", "作业要求"}
            for code, is_software, names in grouped_rows.values():
                aliases = names
                if len(names) >= 3 and names[1] in shared_suffixes:
                    aliases = [names[0] + names[1], *names[2:]]
                for raw_alias in aliases:
                    # 同时保留表格原文与斜杠拆分后的同义名称。
                    self._add(raw_alias, code, is_software)
                    for alias in self._split_inline(raw_alias):
                        self._add(alias, code, is_software)
        finally:
            workbook.close()

    def match(self, file_name: str) -> AbbreviationMatch:
        targets = tuple(normalize_for_match(name) for name in abbreviation_match_names(file_name))
        normalized_file_name = normalize_for_match(file_name)
        for suffix, code in FIXED_SUFFIX_ABBREVIATIONS:
            if normalized_file_name.endswith(normalize_for_match(suffix)):
                return AbbreviationMatch(
                    alias=suffix,
                    code=code,
                    is_software="软件" in normalized_file_name,
                )
        matches = [
            match
            for key, match in self._aliases.items()
            if key and any(key in target for target in targets)
        ]
        if not matches:
            raise AbbreviationError(f"未能从文件简号.xlsx匹配文件类型：{file_name}")
        return max(matches, key=lambda item: len(normalize_for_match(item.alias)))


@lru_cache
def get_abbreviation_registry() -> AbbreviationRegistry:
    return AbbreviationRegistry(get_settings().abbreviation_file_path)
